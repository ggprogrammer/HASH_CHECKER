from openpyxl import Workbook, load_workbook

class excel_creator:
    row_counter = 2
    table_name = ""
    code_type = ""

    def __init__(self, code_type: str):
        wb = Workbook()

        ws = wb.active

        self.code_type = code_type

        if code_type == "SHA256":
            ws.title = f'Результаты {code_type}'
            ws["A1"] = code_type
            ws["B1"] = "Кол-во угроз"

            self.table_name = f"{code_type}_results.xlsx"

            wb.save(f"{code_type}_results.xlsx")
        else:
            ws.title = f'Результаты {code_type}'
            ws["A1"] = code_type
            ws["B1"] = "SHA256_CODE"
            ws["C1"] = "Кол-во угроз"

            self.table_name = f"{code_type}_results.xlsx"

            wb.save(f"{code_type}_results.xlsx")

    def add_row(self, hash_code: str, hash_code256: str, malicious: str):
        wb = load_workbook(self.table_name)

        ws = wb.active
        
        if self.code_type == "SHA256":

            ws.append([hash_code, malicious])

            wb.save(self.table_name)
        else:
            ws.append([hash_code, hash_code256, malicious])

            wb.save(self.table_name)

# if __name__ == "__main__":
#     table = excel_creator("MD5")
#     table.add_row('2132', '321321321', '0')